from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
FILE_NAME = "Register.xlsx"

# Excel файл байхгүй бол үүсгэнэ
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.title = "Register"
    ws.append([
        "Мерчант дугаар",
        "Овог",
        "Нэр",
        "Утасны дугаар",
        "Хүссэн зээлийн хэмжээ",
        "Скоринг хэмжээ",
        "Батлагдсан"
    ])
    wb.save(FILE_NAME)

@app.route("/", methods=["GET", "POST"])
def index():
    message = ""

    if request.method == "POST":
        wb = load_workbook(FILE_NAME)
        ws = wb.active

        ws.append([
            request.form["merchant"],
            request.form["lastname"],
            request.form["firstname"],
            request.form["phone"],
            request.form["loan"],
            request.form["score"],
            request.form["approved"]
        ])

        wb.save(FILE_NAME)
        message = "✅ Бүртгэл амжилттай хийгдлээ"

    return render_template("index.html", message=message)

if __name__ == "__main__":
    app.run(debug=True)
